# -*- coding: utf-8 -*-
"""
Created on Mon Nov 11 22:08:09 2019

@author: dingz
"""

import openpyxl

xfile = openpyxl.load_workbook('F:/test.xlsx')
sheet = xfile.get_sheet_by_name('Sheet')

index = 1
sum_=0
volume = 0

sheet.cell(row=index, column=1, value=index)
sheet.cell(row=index, column=2, value=sum_)
sheet.cell(row=index, column=3, value=volume)
#def write():
    #rb = xlwt.Workbook()
 #   sheet = rb.add_sheet(u'sheet1',cell_overwrite_ok=True)
xfile.save('F:/test.xlsx') 

